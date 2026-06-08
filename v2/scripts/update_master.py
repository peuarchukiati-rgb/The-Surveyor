#!/usr/bin/env python3
"""
Update master stop CSVs from bkk_shelter_list.xlsx.
Preserves folder_url from existing CSVs. Leaves TNS untouched.
After updating per-TOR CSVs, regenerates all_stops.csv.
"""

import csv
import re
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

XLSX_PATH = Path("/Users/peakeuarchukiati/Desktop/the-surveyor/from Mon/bkk_shelter_list_verified.xlsx")
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
SCRIPTS_DIR = Path(__file__).resolve().parent

TORS = ["NS", "SS", "ES", "MS"]
CSV_FIELDS = ["no", "code", "district", "road", "location", "type", "lat", "long", "folder_url"]

# xlsx header → csv field mapping
XLSX_COL_MAP = {
    "ลำดับ": "no",
    "รหัส": "code",
    "เขต": "district",
    "ถนน": "road",
    "ตำแหน่ง": "location",
    "ประเภท": "type",
    "lat": "lat",
    "long": "long",
}


def add_space_to_code(code: str) -> str:
    """Convert 'NS001' → 'NS 001' to match existing CSV format."""
    m = re.match(r"^([A-Z]+)(\d+)$", code)
    if m:
        return f"{m.group(1)} {m.group(2)}"
    return code


def read_xlsx_sheet(wb, sheet_name: str) -> list[dict]:
    """Read a sheet and return list of row dicts with CSV field names."""
    ws = wb[sheet_name]
    headers = [c.value for c in ws[2]]
    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        d = {}
        for i, hdr in enumerate(headers):
            field = XLSX_COL_MAP.get(hdr)
            if field:
                val = str(row[i]).strip() if row[i] is not None else ""
                if field == "code":
                    val = add_space_to_code(val)
                d[field] = val
        d["folder_url"] = ""
        rows.append(d)
    return rows


def read_existing_csv(tor: str) -> dict[str, str]:
    """Read existing CSV and return code → folder_url mapping."""
    csv_path = DATA_DIR / f"stops_{tor}.csv"
    mapping = {}
    if not csv_path.exists():
        return mapping
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            code = row.get("code", "").strip()
            url = row.get("folder_url", "").strip()
            if code:
                mapping[code] = url
    return mapping


def write_csv(tor: str, rows: list[dict]):
    """Write updated CSV."""
    csv_path = DATA_DIR / f"stops_{tor}.csv"
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"xlsx not found: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"Loaded: {XLSX_PATH.name}")
    print(f"Sheets: {wb.sheetnames}\n")

    for tor in TORS:
        # Read new data from xlsx
        new_rows = read_xlsx_sheet(wb, tor)

        # Read existing folder_url mapping
        url_map = read_existing_csv(tor)
        old_codes = set(url_map.keys())
        new_codes = set(r["code"] for r in new_rows)

        # Preserve folder_urls
        for row in new_rows:
            row["folder_url"] = url_map.get(row["code"], "")

        # Diff stats
        added = new_codes - old_codes
        removed = old_codes - new_codes
        kept = old_codes & new_codes

        print(f"--- {tor} ---")
        print(f"  Before: {len(old_codes)} stops")
        print(f"  After:  {len(new_codes)} stops")
        print(f"  Added:  {len(added)}")
        if added:
            for c in sorted(added):
                print(f"    + {c}")
        print(f"  Removed: {len(removed)}")
        if removed:
            for c in sorted(removed):
                print(f"    - {c}")
        print(f"  Kept:   {len(kept)}")

        # Write updated CSV
        write_csv(tor, new_rows)
        print(f"  Wrote: stops_{tor}.csv ({len(new_rows)} rows)\n")

    wb.close()

    # Leave TNS untouched
    tns_path = DATA_DIR / "stops_TNS.csv"
    with open(tns_path, encoding="utf-8") as f:
        tns_count = sum(1 for _ in csv.DictReader(f))
    print(f"--- TNS --- (untouched, {tns_count} rows)\n")

    # Regenerate all_stops.csv
    print("Regenerating all_stops.csv ...")
    result = subprocess.run(
        [sys.executable, str(SCRIPTS_DIR / "build_all_stops.py")],
        capture_output=True, text=True
    )
    print(result.stdout)
    if result.returncode != 0:
        print(f"ERROR: {result.stderr}")


if __name__ == "__main__":
    main()
